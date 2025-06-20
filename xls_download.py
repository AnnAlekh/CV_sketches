import os
import shutil
import urllib.request
from openpyxl import load_workbook

def process_images(file_path, output_folder='TEST_IMAGES_2'):
    # Создаем основную папку
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Папка для ошибок
    error_folder = os.path.join(output_folder, 'errors')
    if not os.path.exists(error_folder):
        os.makedirs(error_folder)
    
    # Загружаем Excel файл
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    # Счетчики
    total = 0
    success = 0
    failed = 0
    error_files = 0
    
    # Проходим по всем строкам (начиная со 2-й)
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[2]  # URL изображения
        angle = row[3]  # Распознанный ракурс
        result = row[4]  # Результат ("да"/"нет"/"?")
        
        if not url or not isinstance(url, str):
            continue
            
        total += 1
        
        try:
            # Генерируем имя файла
            if 'maxposter.ru' in url:
                clean_url = url.split('?')[0]
                filename = clean_url.split('/')[-1]
            elif 'yandexcloud.net' in url or 'media.cm.expert' in url:
                parts = url.split('/')
                filename = f"{parts[-2]}_{parts[-1]}.jpg"
            else:
                filename = url.split('/')[-1].split('?')[0]
            
            # Полный путь к файлу
            full_path = os.path.join(output_folder, filename)
            
            # Скачиваем изображение
            urllib.request.urlretrieve(url, full_path)
            
            # Если результат "нет" - перемещаем в папку ошибок
            if str(result).lower() == 'нет' and angle:
                # Создаем безопасное имя папки (удаляем запрещенные символы)
                safe_angle = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in angle)
                angle_folder = os.path.join(error_folder, safe_angle)
                
                if not os.path.exists(angle_folder):
                    os.makedirs(angle_folder)
                
                # Перемещаем файл
                shutil.move(full_path, os.path.join(angle_folder, filename))
                error_files += 1
                print(f"Ошибка: {filename} -> {safe_angle}")
            else:
                print(f"Успешно: {filename}")
                success += 1
                
        except Exception as e:
            print(f"Ошибка при обработке {url}: {str(e)}")
            failed += 1
    
    # Выводим статистику
    print(f"\nИтого: {total} изображений")
    print(f"Успешно загружено: {success}")
    print(f"Ошибки (перемещены): {error_files}")
    print(f"Не удалось загрузить: {failed}")

# Использование
process_images('/home/ann/Загрузки/Telegram Desktop/recognize-test_2.xlsx')